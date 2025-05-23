import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from date_utils import get_safe_range_name
from datetime import datetime

EXPORT_DIR = "exports"

def export_to_csv(logs, user, date_range, date_from_str):
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)
    safe_user = user.replace("@", "_at_").replace(".", "_")
    safe_range = get_safe_range_name(date_range, date_from_str)
    filename = f"worklogs_{safe_user}_{safe_range}.csv"
    filepath = os.path.join(EXPORT_DIR, filename)

    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["started", "issue", "timeSpent", "comment", "author"])
        writer.writeheader()
        writer.writerows(logs)

    print(f"✅ Exported {len(logs)} logs to {filepath}")

def parse_time_spent(timestring):
    hours, minutes = 0, 0
    parts = timestring.split()
    for part in parts:
        if part.endswith("d"):
            hours += int(part[:-1])*8
        elif part.endswith("h"):
            hours += int(part[:-1])
        elif part.endswith("m"):
            minutes += int(part[:-1])
    return hours * 60 + minutes

def format_minutes_to_hhmm(total_minutes):
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours} hrs {minutes} min"

def export_to_excel(logs, user, date_range, date_from_str):
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)

    safe_user = user.replace("@", "_at_").replace(".", "_")
    safe_range = get_safe_range_name(date_range, date_from_str)
    filename = f"worklogs_{safe_user}_{safe_range}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Work Logs"

    headers = ["started", "issue", "timeSpent", "comment", "author"]
    ws_main.append(headers)

    total_minutes = 0
    summary_by_issue = defaultdict(int)
    summary_by_day = defaultdict(int)

    for row in logs:
        ws_main.append([row[h] for h in headers])

        minutes = parse_time_spent(row["timeSpent"])
        total_minutes += minutes

        summary_by_issue[row["issue"]] += minutes

        date_key = datetime.strptime(row["started"][:10], "%Y-%m-%d").strftime("%Y-%m-%d")
        summary_by_day[date_key] += minutes

    # Add footer row
    footer_label = "TOTAL TIME"
    footer_text = format_minutes_to_hhmm(total_minutes)
    ws_main.append([footer_label, "", footer_text])

    bold_font = Font(bold=True)

    # Style headers and footer
    for col in range(1, len(headers) + 1):
        ws_main.cell(row=1, column=col).font = bold_font
        ws_main.column_dimensions[get_column_letter(col)].width = 20

    footer_row = ws_main.max_row
    for col in range(1, len(headers) + 1):
        ws_main.cell(row=footer_row, column=col).font = bold_font

    # --- Sheet 2: Summary by Issue ---
    ws_issue = wb.create_sheet(title="Summary by Issue")
    ws_issue.append(["Issue", "Total Time"])
    for issue, minutes in sorted(summary_by_issue.items(), key=lambda x: -x[1]):
        ws_issue.append([issue, format_minutes_to_hhmm(minutes)])
    ws_issue.cell(row=1, column=1).font = bold_font
    ws_issue.cell(row=1, column=2).font = bold_font

    # --- Sheet 3: Summary by Day ---
    ws_day = wb.create_sheet(title="Summary by Day")
    ws_day.append(["Date", "Total Time"])
    for day, minutes in sorted(summary_by_day.items()):
        ws_day.append([day, format_minutes_to_hhmm(minutes)])
    ws_day.cell(row=1, column=1).font = bold_font
    ws_day.cell(row=1, column=2).font = bold_font

    # Auto-size columns for all sheets
    for sheet in [ws_issue, ws_day]:
        for col in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(15, max_length + 2)

    wb.save(filepath)
    print(f"✅ Exported {len(logs)} logs to {filepath}")

